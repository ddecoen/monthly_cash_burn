"""Monthly Cash Burn Tracker — Flask + SQLite backend.

Imports NetSuite Statement of Cash Flows (SCF) Excel files, stores
parsed line items per period, and exposes a JSON API for burn analysis,
quarterly views, and runway estimation.
"""

from __future__ import annotations

import os
import re
import sqlite3
from datetime import datetime

from io import BytesIO

from flask import Flask, g, jsonify, render_template, request, send_file
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# App / config
# ---------------------------------------------------------------------------

app = Flask(__name__)

DATABASE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "cash_burn.db")
UPLOAD_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "uploads")

# ---------------------------------------------------------------------------
# Database helpers
# ---------------------------------------------------------------------------

SCHEMA = """\
CREATE TABLE IF NOT EXISTS periods (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    month INTEGER NOT NULL,
    year INTEGER NOT NULL,
    uploaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    UNIQUE(month, year)
);

CREATE TABLE IF NOT EXISTS line_items (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    period_id INTEGER NOT NULL REFERENCES periods(id) ON DELETE CASCADE,
    section TEXT NOT NULL,
    description TEXT NOT NULL,
    amount REAL DEFAULT 0,
    is_subtotal INTEGER DEFAULT 0,
    row_order INTEGER NOT NULL
);
"""


def get_db() -> sqlite3.Connection:
    """Return a per-request database connection stored on Flask's *g*."""
    if "db" not in g:
        g.db = sqlite3.connect(DATABASE)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA journal_mode=WAL;")
        g.db.execute("PRAGMA foreign_keys=ON;")
    return g.db


@app.teardown_appcontext
def close_db(_exc: BaseException | None = None) -> None:
    db = g.pop("db", None)
    if db is not None:
        db.close()


def init_db() -> None:
    """Create tables if they don't already exist."""
    conn = sqlite3.connect(DATABASE)
    try:
        conn.execute("PRAGMA foreign_keys=ON;")
        conn.executescript(SCHEMA)
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Excel parsing helpers
# ---------------------------------------------------------------------------

# Month name → number lookup.
MONTH_MAP: dict[str, int] = {
    "january": 1,
    "february": 2,
    "march": 3,
    "april": 4,
    "may": 5,
    "june": 6,
    "july": 7,
    "august": 8,
    "september": 9,
    "october": 10,
    "november": 11,
    "december": 12,
}

# Regex to pull "Month DD, YYYY" from the period string (row 5).
_PERIOD_RE = re.compile(
    r"(January|February|March|April|May|June|July|August|September"
    r"|October|November|December)\s+\d{1,2},\s*(\d{4})",
    re.IGNORECASE,
)

# Descriptions that mark section headers (no amount on these rows).
_SECTION_HEADERS: dict[str, str] = {
    "cash flows from operating activities": "operating",
    "cash flows from investing activities": "investing",
    "cash flows from financing activities": "financing",
}

# Descriptions that are subtotals — stored with is_subtotal=1.
_SUBTOTAL_DESCRIPTIONS: set[str] = {
    "net cash used in operating activities",
    "net cash provided by operating activities",
    "net cash used in investing activities",
    "net cash provided by investing activities",
    "net cash used in financing activities",
    "net cash provided by financing activities",
    "net decrease in cash and cash equivalents",
    "net increase in cash and cash equivalents",
    "cash and cash equivalents at beginning of period",
    "cash and cash equivalents at end of period",
}


def _parse_amount(raw: object) -> float:
    """Convert a cell value to a float.

    Handles:
    - ``None`` / empty → 0
    - Plain numbers → float
    - Strings with commas and/or parentheses for negatives
    """
    if raw is None:
        return 0.0
    if isinstance(raw, (int, float)):
        return float(raw)
    text = str(raw).strip()
    if not text or text == "-":
        return 0.0

    negative = False
    if text.startswith("(") and text.endswith(")"):
        negative = True
        text = text[1:-1]

    text = text.replace(",", "").replace("$", "").strip()
    try:
        value = float(text)
    except ValueError:
        return 0.0
    return -value if negative else value


def _parse_period_string(text: str) -> tuple[int, int]:
    """Extract (month, year) from a NetSuite period string.

    Examples:
        "One Month Ended December 31, 2025" → (12, 2025)
        "One Month Ended January 31, 2026"  → (1, 2026)

    Raises ``ValueError`` when the string cannot be parsed.
    """
    match = _PERIOD_RE.search(text)
    if not match:
        raise ValueError(f"Cannot parse period from: {text!r}")
    month_name = match.group(1).lower()
    year = int(match.group(2))
    month = MONTH_MAP[month_name]
    return month, year


def _is_section_header(desc: str) -> str | None:
    """Return the section key if *desc* is a section header, else ``None``."""
    normalised = desc.strip().rstrip(":").lower()
    return _SECTION_HEADERS.get(normalised)


def _is_subsection(desc: str) -> bool:
    """Return True if the description is a subsection label (no amount)."""
    lower = desc.strip().rstrip(":").lower()
    return lower in (
        "adjustments to reconcile net loss to cash from operating activities",
        "adjustments to reconcile net income to cash from operating activities",
        "changes in operating assets and liabilities",
    )


def parse_scf_excel(filepath: str) -> dict:
    """Parse a NetSuite Statement of Cash Flows .xlsx file.

    Returns a dict with keys ``month``, ``year``, and ``line_items``
    (a list of dicts with ``section``, ``description``, ``amount``,
    ``is_subtotal``, ``row_order``).
    """
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    # Row 5 contains the period string.
    period_cell = ws.cell(row=5, column=1).value
    if period_cell is None:
        raise ValueError("Row 5 (period string) is empty.")
    month, year = _parse_period_string(str(period_cell))

    # Walk data rows starting at row 8.
    current_section = "operating"
    line_items: list[dict] = []
    row_order = 0

    for row in ws.iter_rows(min_row=8, max_col=2, values_only=True):
        desc_raw = row[0]
        if desc_raw is None:
            continue
        desc = str(desc_raw).strip()
        if not desc:
            continue

        # Check for section header transitions.
        section_key = _is_section_header(desc)
        if section_key is not None:
            current_section = section_key
            continue

        # Skip subsection labels.
        if _is_subsection(desc):
            continue

        amount = _parse_amount(row[1])

        # Rows after the three main sections are "summary" items.
        desc_lower = desc.lower()
        section = current_section
        if desc_lower in (
            "net decrease in cash and cash equivalents",
            "net increase in cash and cash equivalents",
            "cash and cash equivalents at beginning of period",
            "cash and cash equivalents at end of period",
        ):
            section = "summary"

        is_subtotal = 1 if desc_lower in _SUBTOTAL_DESCRIPTIONS else 0

        row_order += 1
        line_items.append(
            {
                "section": section,
                "description": desc,
                "amount": amount,
                "is_subtotal": is_subtotal,
                "row_order": row_order,
            }
        )

    wb.close()
    return {"month": month, "year": year, "line_items": line_items}


# ---------------------------------------------------------------------------
# Database storage helpers
# ---------------------------------------------------------------------------


def _store_period(
    db: sqlite3.Connection,
    month: int,
    year: int,
    line_items: list[dict],
) -> int:
    """Insert (or replace) a period and its line items.

    If the month/year already exists the old data is deleted first.
    Returns the ``period_id``.
    """
    # Delete existing period for this month/year (cascade removes items).
    existing = db.execute(
        "SELECT id FROM periods WHERE month = ? AND year = ?",
        (month, year),
    ).fetchone()
    if existing is not None:
        db.execute("DELETE FROM periods WHERE id = ?", (existing["id"],))

    cursor = db.execute(
        "INSERT INTO periods (month, year) VALUES (?, ?)",
        (month, year),
    )
    period_id = cursor.lastrowid

    for item in line_items:
        db.execute(
            """
            INSERT INTO line_items
                (period_id, section, description, amount, is_subtotal, row_order)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (
                period_id,
                item["section"],
                item["description"],
                item["amount"],
                item["is_subtotal"],
                item["row_order"],
            ),
        )

    db.commit()
    return period_id


# ---------------------------------------------------------------------------
# Serialisation helpers
# ---------------------------------------------------------------------------

_MONTH_NAMES = [
    "",
    "January",
    "February",
    "March",
    "April",
    "May",
    "June",
    "July",
    "August",
    "September",
    "October",
    "November",
    "December",
]

_MONTH_ABBREVS = [
    "", "Jan", "Feb", "Mar", "Apr", "May", "Jun",
    "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
]

_QUARTER_MONTH_RANGES: dict[int, list[int]] = {
    1: [1, 2, 3],
    2: [4, 5, 6],
    3: [7, 8, 9],
    4: [10, 11, 12],
}


def _period_to_dict(row: sqlite3.Row) -> dict:
    d = dict(row)
    d["month_name"] = _MONTH_NAMES[d["month"]]
    return d


def _line_item_to_dict(row: sqlite3.Row) -> dict:
    return dict(row)


# ---------------------------------------------------------------------------
# Routes — pages
# ---------------------------------------------------------------------------


@app.route("/")
def index():
    """Serve the main single-page HTML interface."""
    return render_template("index.html")


# ---------------------------------------------------------------------------
# Routes — REST API
# ---------------------------------------------------------------------------


@app.route("/api/upload", methods=["POST"])
def upload_file():
    """Accept an Excel (.xlsx) upload, parse it, and store the data."""
    if "file" not in request.files:
        return jsonify({"error": "No file part in request."}), 400

    file = request.files["file"]
    if file.filename == "" or file.filename is None:
        return jsonify({"error": "No file selected."}), 400

    if not file.filename.lower().endswith(".xlsx"):
        return jsonify({"error": "Only .xlsx files are accepted."}), 400

    # Ensure upload directory exists.
    os.makedirs(UPLOAD_DIR, exist_ok=True)

    # Save with a timestamped name to avoid collisions.
    safe_name = re.sub(r"[^\w.\-]", "_", file.filename)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    save_path = os.path.join(UPLOAD_DIR, f"{ts}_{safe_name}")
    file.save(save_path)

    try:
        parsed = parse_scf_excel(save_path)
    except Exception as exc:
        return jsonify({"error": f"Failed to parse file: {exc}"}), 400

    db = get_db()
    period_id = _store_period(
        db,
        parsed["month"],
        parsed["year"],
        parsed["line_items"],
    )

    return jsonify(
        {
            "period_id": period_id,
            "month": parsed["month"],
            "year": parsed["year"],
            "month_name": _MONTH_NAMES[parsed["month"]],
            "line_items": parsed["line_items"],
        }
    ), 201


@app.route("/api/periods", methods=["GET"])
def list_periods():
    """Return all periods sorted by year then month."""
    db = get_db()
    rows = db.execute(
        "SELECT * FROM periods ORDER BY year ASC, month ASC"
    ).fetchall()
    return jsonify([_period_to_dict(r) for r in rows])


@app.route("/api/period/<int:period_id>", methods=["GET"])
def get_period(period_id: int):
    """Return a single period with all its line items."""
    db = get_db()
    period = db.execute(
        "SELECT * FROM periods WHERE id = ?", (period_id,)
    ).fetchone()
    if period is None:
        return jsonify({"error": f"Period {period_id} not found."}), 404

    items = db.execute(
        "SELECT * FROM line_items WHERE period_id = ? ORDER BY row_order ASC",
        (period_id,),
    ).fetchall()

    result = _period_to_dict(period)
    result["line_items"] = [_line_item_to_dict(i) for i in items]
    return jsonify(result)


@app.route("/api/period/<int:period_id>", methods=["DELETE"])
def delete_period(period_id: int):
    """Delete a period and its line items."""
    db = get_db()
    period = db.execute(
        "SELECT id FROM periods WHERE id = ?", (period_id,)
    ).fetchone()
    if period is None:
        return jsonify({"error": f"Period {period_id} not found."}), 404

    db.execute("DELETE FROM periods WHERE id = ?", (period_id,))
    db.commit()
    return jsonify({"message": f"Period {period_id} deleted."}), 200


def _build_quarterly_data(year: int, quarter: int) -> dict:
    """Build quarterly SCF data. Returns the same structure as the JSON endpoint."""
    months = _QUARTER_MONTH_RANGES[quarter]

    db = get_db()

    # Fetch periods for the requested quarter.
    placeholders = ",".join("?" for _ in months)
    periods = db.execute(
        f"SELECT * FROM periods WHERE year = ? AND month IN ({placeholders}) "
        "ORDER BY month ASC",
        [year] + months,
    ).fetchall()

    period_map: dict[int, sqlite3.Row] = {p["month"]: p for p in periods}

    # Build a stable ordered list of descriptions from whichever periods
    # exist, preserving row_order.
    description_order: list[str] = []
    seen_descriptions: set[str] = set()
    section_map: dict[str, str] = {}
    subtotal_map: dict[str, int] = {}

    for m in months:
        if m not in period_map:
            continue
        items = db.execute(
            "SELECT * FROM line_items WHERE period_id = ? ORDER BY row_order ASC",
            (period_map[m]["id"],),
        ).fetchall()
        for item in items:
            desc = item["description"]
            if desc not in seen_descriptions:
                description_order.append(desc)
                seen_descriptions.add(desc)
                section_map[desc] = item["section"]
                subtotal_map[desc] = item["is_subtotal"]

    # Build month metadata the frontend expects.
    months_meta = [
        {
            "year": year,
            "month": m,
            "available": m in period_map,
            "label": f"{_MONTH_NAMES[m]} {year}",
        }
        for m in months
    ]

    # Map (section, is_subtotal) to the row type the frontend uses.
    section_header_descs = {
        "cash flows from operating activities",
        "cash flows from investing activities",
        "cash flows from financing activities",
    }
    subsection_descs = {
        "adjustments to reconcile net loss to cash from operating activities",
        "adjustments to reconcile net income to cash from operating activities",
        "changes in operating assets and liabilities",
    }

    def _row_type(desc: str, section: str, is_sub: int) -> str:
        dl = desc.strip().rstrip(":").lower()
        if dl in section_header_descs:
            return "section_header"
        if dl in subsection_descs:
            return "subsection"
        if section == "summary":
            return "grand_total"
        if is_sub:
            return "subtotal"
        return "line"

    # Insert section headers and subsections into the row list so the
    # frontend can render them.  Walk the ordered descriptions and
    # emit synthetic header rows when the section changes.
    section_titles = {
        "operating": "Cash flows from operating activities",
        "investing": "Cash flows from investing activities",
        "financing": "Cash flows from financing activities",
    }

    rows: list[dict] = []
    prev_section: str | None = None
    emitted_adjustments = False
    emitted_changes = False

    for desc in description_order:
        section = section_map[desc]
        is_sub = subtotal_map[desc]

        # Emit section header on section transition.
        if section != prev_section and section in section_titles:
            rows.append(
                {
                    "description": section_titles[section],
                    "type": "section_header",
                    "values": [None] * len(months),
                }
            )
            prev_section = section
            emitted_adjustments = False
            emitted_changes = False

        # Emit subsection headers within operating.
        if section == "operating" and not is_sub:
            dl = desc.lower()
            if (
                not emitted_adjustments
                and dl == "depreciation and amortization expense"
            ):
                rows.append(
                    {
                        "description": "Adjustments to reconcile net loss to cash from operating activities:",
                        "type": "subsection",
                        "values": [None] * len(months),
                    }
                )
                emitted_adjustments = True
            if not emitted_changes and dl == "accounts receivable":
                rows.append(
                    {
                        "description": "Changes in operating assets and liabilities:",
                        "type": "subsection",
                        "values": [None] * len(months),
                    }
                )
                emitted_changes = True

        # Build per-month values (None when month not uploaded).
        values: list[float | None] = []
        for m in months:
            if m not in period_map:
                values.append(None)
            else:
                row = db.execute(
                    "SELECT amount FROM line_items "
                    "WHERE period_id = ? AND description = ?",
                    (period_map[m]["id"], desc),
                ).fetchone()
                values.append(row["amount"] if row else 0.0)

        available_vals = [v for v in values if v is not None]
        total = round(sum(available_vals), 2) if available_vals else None

        rows.append(
            {
                "description": desc,
                "type": _row_type(desc, section, is_sub),
                "values": values,
                "total": total,
            }
        )

    return {
        "year": year,
        "quarter": quarter,
        "months": months_meta,
        "rows": rows,
    }


@app.route("/api/quarterly", methods=["GET"])
def quarterly_view():
    """Return a quarterly view of line items for column display.

    Query params:
        year   — required, e.g. 2025
        quarter — required, 1-4 (Q1=Jan-Mar … Q4=Oct-Dec)

    Returns each line item description mapped to an array of monthly
    values plus a "Total" sum column.
    """
    year_raw = request.args.get("year")
    quarter_raw = request.args.get("quarter")
    if year_raw is None or quarter_raw is None:
        return jsonify({"error": "'year' and 'quarter' are required."}), 400

    try:
        year = int(year_raw)
        quarter = int(quarter_raw)
    except (TypeError, ValueError):
        return jsonify({"error": "'year' and 'quarter' must be integers."}), 400

    if quarter not in (1, 2, 3, 4):
        return jsonify({"error": "'quarter' must be 1, 2, 3, or 4."}), 400

    return jsonify(_build_quarterly_data(year, quarter))


# ---------------------------------------------------------------------------
# Excel export — investor-friendly burn report
# ---------------------------------------------------------------------------


def _extract_row_value(rows: list[dict], desc_fragment: str) -> list:
    """Find a row whose description contains *desc_fragment* (case-insensitive)
    and return its values list.  Prefers exact matches over substring.
    Returns a list of Nones if not found."""
    frag = desc_fragment.lower()
    # First pass: exact match on the stripped, lowered description.
    for r in rows:
        d = r.get("description", "").strip().lower()
        if d == frag or d == frag.rstrip(":"):
            return r.get("values", [])
    # Second pass: starts-with match.
    for r in rows:
        d = r.get("description", "").strip().lower()
        if d.startswith(frag):
            return r.get("values", [])
    # Third pass: substring match.
    for r in rows:
        if frag in r.get("description", "").lower():
            return r.get("values", [])
    return []


def _safe(val: float | None) -> float | None:
    """Return the value unchanged, or None."""
    if val is None:
        return None
    return round(val, 2)


def _sum_available(vals: list[float | None]) -> float | None:
    """Sum only non-None values; return None if all are None."""
    nums = [v for v in vals if v is not None]
    return round(sum(nums), 2) if nums else None


def _last_available(vals: list[float | None]) -> float | None:
    """Return the last non-None value."""
    for v in reversed(vals):
        if v is not None:
            return v
    return None


@app.route("/api/export", methods=["GET"])
def export_excel():
    """Export an investor-friendly monthly cash burn report as Excel.

    Two sheets:
      1. **Cash Burn Report** — high-level summary, key metrics, and
         simplified monthly/quarterly cash-flow view.
      2. **Detailed Cash Flows** — the full SCF line items.

    Query params:
        year    — required, e.g. 2025
        quarter — required, 1-4
    """
    year_raw = request.args.get("year")
    quarter_raw = request.args.get("quarter")
    if year_raw is None or quarter_raw is None:
        return jsonify({"error": "'year' and 'quarter' are required."}), 400

    try:
        year = int(year_raw)
        quarter = int(quarter_raw)
    except (TypeError, ValueError):
        return jsonify({"error": "'year' and 'quarter' must be integers."}), 400

    if quarter not in (1, 2, 3, 4):
        return jsonify({"error": "'quarter' must be 1, 2, 3, or 4."}), 400

    qdata = _build_quarterly_data(year, quarter)
    months_meta = qdata["months"]
    all_rows = qdata["rows"]

    # Convenience: month column headers.
    month_headers = [
        f"{_MONTH_ABBREVS[m['month']]} {m['year']}" for m in months_meta
    ]
    num_months = len(months_meta)
    # Columns: Description | month1 | month2 | month3 | Quarter Total
    num_cols = 1 + num_months + 1

    # ── Extract key line-item values from the quarterly data ──────────
    v_net_income = _extract_row_value(all_rows, "net (loss) income")
    v_da = _extract_row_value(all_rows, "depreciation and amortization")
    v_ar = _extract_row_value(all_rows, "accounts receivable")
    v_prepaids = _extract_row_value(all_rows, "prepaid expenses")
    v_other_assets = _extract_row_value(all_rows, "other assets")
    v_ap = _extract_row_value(all_rows, "accounts payable")
    v_accrued = _extract_row_value(all_rows, "accrued expenses")
    v_deferred = _extract_row_value(all_rows, "deferred revenue")
    v_op_total = _extract_row_value(all_rows, "net cash used in operating")
    if not any(v is not None for v in v_op_total):
        v_op_total = _extract_row_value(
            all_rows, "net cash provided by operating"
        )
    v_capex = _extract_row_value(all_rows, "purchases of property")
    v_inv_total = _extract_row_value(all_rows, "net cash used in investing")
    if not any(v is not None for v in v_inv_total):
        v_inv_total = _extract_row_value(
            all_rows, "net cash provided by investing"
        )
    v_stock = _extract_row_value(all_rows, "proceeds from stock")
    v_equity = _extract_row_value(all_rows, "other equity")
    v_fin_total = _extract_row_value(
        all_rows, "net cash provided by financing"
    )
    if not any(v is not None for v in v_fin_total):
        v_fin_total = _extract_row_value(
            all_rows, "net cash used in financing"
        )
    v_net_change = _extract_row_value(all_rows, "net decrease in cash")
    if not any(v is not None for v in v_net_change):
        v_net_change = _extract_row_value(all_rows, "net increase in cash")
    v_begin_cash = _extract_row_value(all_rows, "beginning of period")
    v_end_cash = _extract_row_value(all_rows, "end of period")

    # Working capital = sum of AR + prepaids + other assets + AP +
    # accrued + deferred.
    v_wc = []
    for i in range(num_months):
        components = [
            v_ar[i] if i < len(v_ar) else None,
            v_prepaids[i] if i < len(v_prepaids) else None,
            v_other_assets[i] if i < len(v_other_assets) else None,
            v_ap[i] if i < len(v_ap) else None,
            v_accrued[i] if i < len(v_accrued) else None,
            v_deferred[i] if i < len(v_deferred) else None,
        ]
        non_none = [c for c in components if c is not None]
        v_wc.append(round(sum(non_none), 2) if non_none else None)

    # ── Compute key metrics ───────────────────────────────────────────
    net_changes = [v for v in v_net_change if v is not None]
    avg_monthly_burn = (
        round(-sum(net_changes) / len(net_changes), 2)
        if net_changes
        else 0
    )
    op_totals = [v for v in v_op_total if v is not None]
    avg_op_burn = (
        round(-sum(op_totals) / len(op_totals), 2) if op_totals else 0
    )
    ending_cash = _last_available(v_end_cash)
    beginning_cash_first = next(
        (v for v in v_begin_cash if v is not None), None
    )
    qtr_net_change = _sum_available(v_net_change)
    runway = (
        round(ending_cash / avg_monthly_burn, 1)
        if ending_cash and avg_monthly_burn > 0
        else None
    )

    # ══════════════════════════════════════════════════════════════════
    # Build workbook
    # ══════════════════════════════════════════════════════════════════
    wb = Workbook()

    # ── Reusable styles ───────────────────────────────────────────────
    num_fmt = '#,##0;(#,##0);"-"'
    pct_fmt = '0.0%'
    font_title = Font(bold=True, size=16, color="1A3C6E")
    font_subtitle = Font(bold=True, size=12, color="333333")
    font_section = Font(bold=True, size=11, color="1A3C6E")
    font_bold = Font(bold=True)
    font_bold_white = Font(bold=True, color="FFFFFF")
    font_italic = Font(italic=True, color="666666")
    font_normal = Font(size=10)
    font_metric_label = Font(size=10, color="5F6368")
    font_metric_val = Font(bold=True, size=14)
    fill_header = PatternFill(
        start_color="1A3C6E", end_color="1A3C6E", fill_type="solid"
    )
    fill_light_blue = PatternFill(
        start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"
    )
    fill_section = PatternFill(
        start_color="E8EDF2", end_color="E8EDF2", fill_type="solid"
    )
    fill_subtotal = PatternFill(
        start_color="F0F4F8", end_color="F0F4F8", fill_type="solid"
    )
    fill_metric = PatternFill(
        start_color="F5F7FA", end_color="F5F7FA", fill_type="solid"
    )
    fill_green = PatternFill(
        start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"
    )
    fill_red = PatternFill(
        start_color="FFEBEE", end_color="FFEBEE", fill_type="solid"
    )
    border_bottom = Border(bottom=Side(style="thin"))
    border_top = Border(top=Side(style="thin"))
    border_top_bottom = Border(
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    border_double = Border(
        top=Side(style="thin"), bottom=Side(style="double")
    )
    align_right = Alignment(horizontal="right")
    align_center = Alignment(horizontal="center")
    align_left = Alignment(horizontal="left", indent=1)
    align_left_2 = Alignment(horizontal="left", indent=2)

    def _wv(cell, val, fmt=num_fmt):
        """Write value: number or em-dash for None."""
        if val is None:
            cell.value = "\u2014"
            cell.alignment = align_center
            cell.font = Font(color="999999")
        else:
            cell.value = val
            cell.number_format = fmt
            cell.alignment = align_right

    def _write_data_row(ws, r, desc, vals, total, bold=False,
                        fill=None, indent=False, border=None):
        """Write a full row with description + month values + total."""
        dc = ws.cell(row=r, column=1, value=desc)
        dc.font = font_bold if bold else font_normal
        if indent:
            dc.alignment = align_left
        if fill:
            dc.fill = fill
        for i, v in enumerate(vals):
            c = ws.cell(row=r, column=2 + i)
            _wv(c, v)
            if bold:
                c.font = font_bold
            if fill:
                c.fill = fill
            if border:
                c.border = border
        tc = ws.cell(row=r, column=num_cols)
        _wv(tc, total)
        if bold:
            tc.font = font_bold
        if fill:
            tc.fill = fill
        if border:
            tc.border = border

    def _set_col_widths(ws):
        ws.column_dimensions["A"].width = 42
        for ci in range(2, num_cols + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 18

    # ══════════════════════════════════════════════════════════════════
    # SHEET 1: Cash Burn Report
    # ══════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Cash Burn Report"
    _set_col_widths(ws1)
    ws1.sheet_properties.tabColor = "1A73E8"

    first_m = months_meta[0]["month"]
    last_m = months_meta[-1]["month"]
    period_label = (
        f"Q{quarter} {year} "
        f"({_MONTH_ABBREVS[first_m]} \u2013 {_MONTH_ABBREVS[last_m]})"
    )

    # ── Title block ───────────────────────────────────────────────────
    r = 1
    ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=num_cols)
    c = ws1.cell(row=r, column=1, value="Monthly Cash Burn Report")
    c.font = font_title
    r = 2
    ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=num_cols)
    c = ws1.cell(row=r, column=1, value=period_label)
    c.font = font_subtitle
    r = 3
    ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=num_cols)
    c = ws1.cell(row=r, column=1, value="Amounts in thousands")
    c.font = font_italic

    # ── Key Metrics bar (row 5-6) ─────────────────────────────────────
    r = 5
    ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=num_cols)
    c = ws1.cell(row=r, column=1, value="KEY METRICS")
    c.font = font_section
    c.fill = fill_section
    for ci in range(2, num_cols + 1):
        ws1.cell(row=r, column=ci).fill = fill_section

    # Metric labels on row 6, values on row 7.
    metric_defs = [
        ("Avg Monthly Burn", avg_monthly_burn, fill_red if avg_monthly_burn > 0 else fill_green),
        ("Avg Operating Burn", avg_op_burn, fill_red if avg_op_burn > 0 else fill_green),
        ("Ending Cash Position", ending_cash, fill_metric),
        ("Estimated Runway", None, fill_metric),  # handled specially
        ("Qtr Net Cash Change", qtr_net_change, fill_red if (qtr_net_change or 0) < 0 else fill_green),
    ]
    r = 6
    for i, (label, val, bg) in enumerate(metric_defs):
        col = i + 1
        lc = ws1.cell(row=r, column=col, value=label)
        lc.font = font_metric_label
        lc.fill = bg
        vc = ws1.cell(row=r + 1, column=col)
        vc.fill = bg
        if label == "Estimated Runway":
            if runway is not None:
                vc.value = f"{runway} months"
            else:
                vc.value = "N/A (not burning)"
            vc.font = font_metric_val
        else:
            _wv(vc, val)
            vc.font = font_metric_val

    # ── Cash Position Summary (rows 9+) ──────────────────────────────
    r = 9
    ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=num_cols)
    c = ws1.cell(row=r, column=1, value="CASH POSITION")
    c.font = font_section
    c.fill = fill_section
    for ci in range(2, num_cols + 1):
        ws1.cell(row=r, column=ci).fill = fill_section

    # Column headers.
    r = 10
    headers = [""] + month_headers + ["Quarter"]
    for ci, h in enumerate(headers, 1):
        c = ws1.cell(row=r, column=ci, value=h)
        c.font = font_bold
        c.border = border_bottom
        c.fill = fill_light_blue
        if ci > 1:
            c.alignment = align_center

    r = 11
    _write_data_row(ws1, r, "Beginning Cash", v_begin_cash,
                    beginning_cash_first, indent=True)
    r = 12
    _write_data_row(ws1, r, "Net Cash Change", v_net_change,
                    _sum_available(v_net_change), indent=True)
    r = 13
    _write_data_row(ws1, r, "Ending Cash", v_end_cash,
                    _last_available(v_end_cash), bold=True,
                    border=border_double)

    # ── Cash Flow Summary (rows 15+) ─────────────────────────────────
    r = 15
    ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=num_cols)
    c = ws1.cell(row=r, column=1, value="CASH FLOW SUMMARY")
    c.font = font_section
    c.fill = fill_section
    for ci in range(2, num_cols + 1):
        ws1.cell(row=r, column=ci).fill = fill_section

    r = 16
    headers2 = [""] + month_headers + ["Quarter Total"]
    for ci, h in enumerate(headers2, 1):
        c = ws1.cell(row=r, column=ci, value=h)
        c.font = font_bold
        c.border = border_bottom
        c.fill = fill_light_blue
        if ci > 1:
            c.alignment = align_center

    r = 17
    _write_data_row(ws1, r, "Cash from Operations", v_op_total,
                    _sum_available(v_op_total), indent=True)
    r = 18
    _write_data_row(ws1, r, "Cash from Investing", v_inv_total,
                    _sum_available(v_inv_total), indent=True)
    r = 19
    _write_data_row(ws1, r, "Cash from Financing", v_fin_total,
                    _sum_available(v_fin_total), indent=True)
    r = 20
    _write_data_row(ws1, r, "Net Cash Flow", v_net_change,
                    _sum_available(v_net_change), bold=True,
                    fill=fill_subtotal, border=border_top)

    # ── Operating Detail (rows 22+) ──────────────────────────────────
    r = 22
    ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=num_cols)
    c = ws1.cell(row=r, column=1, value="OPERATING CASH FLOW DETAIL")
    c.font = font_section
    c.fill = fill_section
    for ci in range(2, num_cols + 1):
        ws1.cell(row=r, column=ci).fill = fill_section

    r = 23
    for ci, h in enumerate(headers2, 1):
        c = ws1.cell(row=r, column=ci, value=h)
        c.font = font_bold
        c.border = border_bottom
        c.fill = fill_light_blue
        if ci > 1:
            c.alignment = align_center

    r = 24
    _write_data_row(ws1, r, "Net Income", v_net_income,
                    _sum_available(v_net_income), indent=True)
    r = 25
    _write_data_row(ws1, r, "Depreciation & Amortization", v_da,
                    _sum_available(v_da), indent=True)
    r = 26
    _write_data_row(ws1, r, "Working Capital Changes", v_wc,
                    _sum_available(v_wc), indent=True)

    # Working capital breakdown (indented further).
    wc_items = [
        ("Accounts Receivable", v_ar),
        ("Prepaid Expenses & Other", v_prepaids),
        ("Other Assets", v_other_assets),
        ("Accounts Payable", v_ap),
        ("Accrued Expenses & Other Liabilities", v_accrued),
        ("Deferred Revenue", v_deferred),
    ]
    r = 27
    for label, vals in wc_items:
        dc = ws1.cell(row=r, column=1, value=label)
        dc.font = Font(size=10, color="666666")
        dc.alignment = align_left_2
        for i, v in enumerate(vals[:num_months]):
            c = ws1.cell(row=r, column=2 + i)
            _wv(c, v)
            c.font = Font(size=10, color="666666")
        tc = ws1.cell(row=r, column=num_cols)
        _wv(tc, _sum_available(vals[:num_months]))
        tc.font = Font(size=10, color="666666")
        r += 1

    _write_data_row(ws1, r, "Net Cash from Operations", v_op_total,
                    _sum_available(v_op_total), bold=True,
                    fill=fill_subtotal, border=border_double)

    # ── Investing & Financing detail ─────────────────────────────────
    r += 2
    ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=num_cols)
    c = ws1.cell(row=r, column=1, value="INVESTING & FINANCING DETAIL")
    c.font = font_section
    c.fill = fill_section
    for ci in range(2, num_cols + 1):
        ws1.cell(row=r, column=ci).fill = fill_section

    r += 1
    for ci, h in enumerate(headers2, 1):
        c = ws1.cell(row=r, column=ci, value=h)
        c.font = font_bold
        c.border = border_bottom
        c.fill = fill_light_blue
        if ci > 1:
            c.alignment = align_center

    r += 1
    _write_data_row(ws1, r, "Capital Expenditures", v_capex,
                    _sum_available(v_capex), indent=True)
    r += 1
    _write_data_row(ws1, r, "Net Cash from Investing", v_inv_total,
                    _sum_available(v_inv_total), bold=True,
                    border=border_top)
    r += 1  # blank
    r += 1
    _write_data_row(ws1, r, "Stock Issuance Proceeds", v_stock,
                    _sum_available(v_stock), indent=True)
    r += 1
    _write_data_row(ws1, r, "Other Equity Transactions", v_equity,
                    _sum_available(v_equity), indent=True)
    r += 1
    _write_data_row(ws1, r, "Net Cash from Financing", v_fin_total,
                    _sum_available(v_fin_total), bold=True,
                    border=border_top)

    # ── Confidentiality footer ───────────────────────────────────────
    r += 2
    ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=num_cols)
    c = ws1.cell(
        row=r, column=1,
        value="CONFIDENTIAL — For intended recipients only. "
              "Do not distribute without permission.",
    )
    c.font = Font(italic=True, size=8, color="999999")

    # Print setup.
    ws1.print_title_rows = "1:3"
    ws1.page_setup.orientation = "landscape"
    ws1.page_setup.fitToWidth = 1

    # ══════════════════════════════════════════════════════════════════
    # SHEET 2: Detailed Cash Flows (full SCF)
    # ══════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet(title="Detailed Cash Flows")
    _set_col_widths(ws2)
    ws2.sheet_properties.tabColor = "5F6368"

    r = 1
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=num_cols)
    c = ws2.cell(row=r, column=1, value="Condensed Statement of Cash Flows")
    c.font = Font(bold=True, size=14)
    r = 2
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=num_cols)
    c = ws2.cell(row=r, column=1, value=period_label)
    c.font = Font(italic=True, size=11)
    r = 3
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=num_cols)
    c = ws2.cell(row=r, column=1, value="(amounts in thousands, unaudited)")
    c.font = font_italic

    r = 5
    for ci, h in enumerate(headers2, 1):
        c = ws2.cell(row=r, column=ci, value=h)
        c.font = font_bold_white
        c.fill = fill_header
        c.border = border_bottom
        if ci > 1:
            c.alignment = align_center

    r = 6
    for row_data in all_rows:
        desc = row_data["description"]
        row_type = row_data["type"]
        values = row_data.get("values", [])
        total = row_data.get("total")

        dc = ws2.cell(row=r, column=1)

        if row_type == "section_header":
            dc.value = desc
            dc.font = font_bold
            dc.fill = fill_section
            for ci in range(2, num_cols + 1):
                ws2.cell(row=r, column=ci).fill = fill_section

        elif row_type == "subsection":
            dc.value = f"  {desc}"
            dc.font = Font(italic=True)

        elif row_type == "line":
            dc.value = f"  {desc}"
            for i, val in enumerate(values):
                _wv(ws2.cell(row=r, column=2 + i), val)
            _wv(ws2.cell(row=r, column=num_cols), total)

        elif row_type == "subtotal":
            dc.value = desc
            dc.font = font_bold
            dc.fill = fill_subtotal
            for i, val in enumerate(values):
                c = ws2.cell(row=r, column=2 + i)
                c.fill = fill_subtotal
                c.border = border_top
                _wv(c, val)
                c.font = font_bold
            tc = ws2.cell(row=r, column=num_cols)
            tc.fill = fill_subtotal
            tc.border = border_top
            _wv(tc, total)
            tc.font = font_bold

        elif row_type == "grand_total":
            dc.value = desc
            dc.font = font_bold
            dl = desc.lower()
            for i, val in enumerate(values):
                c = ws2.cell(row=r, column=2 + i)
                c.border = border_double
                _wv(c, val)
                c.font = font_bold
            tc = ws2.cell(row=r, column=num_cols)
            tc.border = border_double
            tc.font = font_bold
            if "beginning" in dl:
                tc.value = "\u2014"
                tc.alignment = align_center
            elif "end" in dl:
                _wv(tc, _last_available(values))
                tc.font = font_bold
            else:
                _wv(tc, total)
                tc.font = font_bold

        r += 1

    # ── Stream the workbook ───────────────────────────────────────────
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Cash_Burn_Report_Q{quarter}_{year}.xlsx"
    return send_file(
        output,
        mimetype=(
            "application/vnd.openxmlformats-officedocument"
            ".spreadsheetml.sheet"
        ),
        as_attachment=True,
        download_name=filename,
    )


@app.route("/api/burn-summary", methods=["GET"])
def burn_summary():
    """Return burn analysis across all stored periods.

    Optional query param ``cash_on_hand`` for runway estimation.
    """
    db = get_db()
    periods = db.execute(
        "SELECT * FROM periods ORDER BY year ASC, month ASC"
    ).fetchall()

    if not periods:
        return jsonify(
            {
                "months_of_data": 0,
                "monthly_burn_trend": [],
                "avg_monthly_burn": 0,
                "avg_net_cash_change": 0,
                "avg_operating_burn": 0,
                "cash_position": None,
                "cash_position_over_time": [],
                "runway_months": None,
            }
        )

    monthly_burn_trend: list[dict] = []
    cash_positions: list[dict] = []
    operating_burns: list[float] = []

    for period in periods:
        pid = period["id"]
        month_name = _MONTH_NAMES[period["month"]]
        label = f"{month_name} {period['year']}"

        # Net cash change = "Net decrease in cash and cash equivalents"
        # or "Net increase …". Fall back to summing the three section
        # subtotals if the summary row is missing.
        net_row = db.execute(
            "SELECT amount FROM line_items "
            "WHERE period_id = ? AND section = 'summary' "
            "AND LOWER(description) LIKE 'net % in cash and cash equivalents'",
            (pid,),
        ).fetchone()

        if net_row is not None:
            net_change = net_row["amount"]
        else:
            # Fallback: sum the three section subtotals.
            subtotals = db.execute(
                "SELECT amount FROM line_items "
                "WHERE period_id = ? AND is_subtotal = 1 "
                "AND section IN ('operating', 'investing', 'financing')",
                (pid,),
            ).fetchall()
            net_change = sum(r["amount"] for r in subtotals)

        # Operating burn = net cash used in operating activities.
        op_row = db.execute(
            "SELECT amount FROM line_items "
            "WHERE period_id = ? AND is_subtotal = 1 "
            "AND section = 'operating'",
            (pid,),
        ).fetchone()
        if op_row is not None:
            operating_burns.append(op_row["amount"])

        monthly_burn_trend.append(
            {
                "period": label,
                "month": period["month"],
                "year": period["year"],
                "net_cash_change": round(net_change, 2),
            }
        )

        # Cash position = ending cash for the period.
        cash_end_row = db.execute(
            "SELECT amount FROM line_items "
            "WHERE period_id = ? "
            "AND LOWER(description) = 'cash and cash equivalents at end of period'",
            (pid,),
        ).fetchone()

        cash_positions.append(
            {
                "period": label,
                "month": period["month"],
                "year": period["year"],
                "cash_position": round(
                    cash_end_row["amount"] if cash_end_row else 0.0, 2
                ),
            }
        )

    # Average monthly burn (negative net_cash_change means burning cash,
    # so average burn is the negated mean).
    total_net = sum(item["net_cash_change"] for item in monthly_burn_trend)
    months_of_data = len(monthly_burn_trend)
    avg_monthly_burn = round(-total_net / months_of_data, 2)
    avg_net_cash_change = round(total_net / months_of_data, 2)
    avg_operating_burn = (
        round(sum(operating_burns) / len(operating_burns), 2)
        if operating_burns
        else 0
    )
    latest_cash = (
        cash_positions[-1]["cash_position"] if cash_positions else None
    )

    # Runway estimate.
    cash_on_hand_raw = request.args.get("cash_on_hand")
    runway_months: float | None = None
    if cash_on_hand_raw is not None:
        try:
            cash_on_hand = float(cash_on_hand_raw)
            if avg_monthly_burn > 0:
                runway_months = round(cash_on_hand / avg_monthly_burn, 1)
        except (TypeError, ValueError):
            pass

    return jsonify(
        {
            "months_of_data": months_of_data,
            "monthly_burn_trend": monthly_burn_trend,
            "avg_monthly_burn": avg_monthly_burn,
            "avg_net_cash_change": avg_net_cash_change,
            "avg_operating_burn": avg_operating_burn,
            "cash_position": latest_cash,
            "cash_position_over_time": cash_positions,
            "runway_months": runway_months,
        }
    )


# ---------------------------------------------------------------------------
# Entrypoint
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    init_db()
    app.run(host="0.0.0.0", port=5000, debug=True)
